import os, re, json, csv, io
from flask import Flask, request, render_template, redirect, url_for, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
import ipaddress

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///network_audit.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ─── Models ───────────────────────────────────────────────────────────────────

class Device(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    hostname    = db.Column(db.String(100))
    vendor      = db.Column(db.String(50))
    filename    = db.Column(db.String(200))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    interfaces  = db.relationship('Interface', backref='device', cascade='all, delete-orphan')
    protocols   = db.relationship('Protocol',  backref='device', cascade='all, delete-orphan')
    acl_rules   = db.relationship('ACLRule',   backref='device', cascade='all, delete-orphan')
    validations = db.relationship('Validation', backref='device', cascade='all, delete-orphan')

class Interface(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey('device.id'))
    name      = db.Column(db.String(100))
    ip        = db.Column(db.String(50))
    mask      = db.Column(db.String(50))
    description = db.Column(db.String(200))

class Protocol(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey('device.id'))
    type      = db.Column(db.String(20))   # OSPF / BGP
    details   = db.Column(db.Text)

class ACLRule(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey('device.id'))
    acl_name  = db.Column(db.String(100))
    rule      = db.Column(db.Text)

class Validation(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    device_id = db.Column(db.Integer, db.ForeignKey('device.id'))
    check     = db.Column(db.String(200))
    status    = db.Column(db.String(10))   # PASS / FAIL
    message   = db.Column(db.Text)

# ─── Parsers ──────────────────────────────────────────────────────────────────

def detect_vendor(content, filename):
    fn = filename.lower()
    if fn.endswith('.conf') or 'juniper' in fn:
        return 'Juniper'
    if 'huawei' in fn or content.strip().startswith('sysname') or 'sysname' in content[:200]:
        return 'Huawei'
    return 'Cisco'

def parse_cisco(content):
    data = {'hostname': '', 'interfaces': [], 'protocols': [], 'acl_rules': []}
    lines = content.splitlines()

    # hostname
    for l in lines:
        m = re.match(r'^hostname\s+(\S+)', l)
        if m:
            data['hostname'] = m.group(1)
            break

    # interfaces
    iface = None
    for l in lines:
        m = re.match(r'^\s*interface\s+(\S+)', l)
        if m:
            if iface: data['interfaces'].append(iface)
            iface = {'name': m.group(1), 'ip': '', 'mask': '', 'description': ''}
        elif iface:
            mi = re.match(r'\s+ip address\s+(\S+)\s+(\S+)', l)
            if mi:
                iface['ip'] = mi.group(1); iface['mask'] = mi.group(2)
            md = re.match(r'\s+description\s+(.*)', l)
            if md:
                iface['description'] = md.group(1).strip()
            if re.match(r'^[a-zA-Z!]', l) and not l.startswith(' '):
                data['interfaces'].append(iface); iface = None
    if iface:
        data['interfaces'].append(iface)

    # OSPF
    in_ospf = False
    ospf_details = {}
    for l in lines:
        if re.match(r'^router ospf', l):
            in_ospf = True
            ospf_details = {'router_id': '', 'networks': []}
        elif in_ospf:
            if re.match(r'^[a-zA-Z!]', l) and not l.startswith(' '):
                in_ospf = False
                data['protocols'].append({'type': 'OSPF', 'details': json.dumps(ospf_details)})
            else:
                m = re.match(r'\s+router-id\s+(\S+)', l)
                if m: ospf_details['router_id'] = m.group(1)
                m = re.match(r'\s+network\s+(\S+)\s+\S+\s+area\s+(\S+)', l)
                if m: ospf_details['networks'].append({'net': m.group(1), 'area': m.group(2)})
    if in_ospf and ospf_details:
        data['protocols'].append({'type': 'OSPF', 'details': json.dumps(ospf_details)})

    # BGP
    in_bgp = False
    bgp_details = {}
    for l in lines:
        if re.match(r'^router bgp', l):
            in_bgp = True
            bgp_details = {'router_id': '', 'neighbors': []}
        elif in_bgp:
            if re.match(r'^[a-zA-Z!]', l) and not l.startswith(' '):
                in_bgp = False
                data['protocols'].append({'type': 'BGP', 'details': json.dumps(bgp_details)})
            else:
                m = re.match(r'\s+bgp router-id\s+(\S+)', l)
                if m: bgp_details['router_id'] = m.group(1)
                m = re.match(r'\s+neighbor\s+(\S+)\s+remote-as\s+(\S+)', l)
                if m: bgp_details['neighbors'].append({'ip': m.group(1), 'as': m.group(2)})
    if in_bgp and bgp_details:
        data['protocols'].append({'type': 'BGP', 'details': json.dumps(bgp_details)})

    # ACLs
    for l in lines:
        m = re.match(r'^access-list\s+(\S+)\s+(.*)', l)
        if m:
            data['acl_rules'].append({'acl_name': f"ACL {m.group(1)}", 'rule': m.group(2).strip()})

    return data

def parse_huawei(content):
    data = {'hostname': '', 'interfaces': [], 'protocols': [], 'acl_rules': []}
    lines = content.splitlines()

    for l in lines:
        m = re.match(r'^sysname\s+(\S+)', l)
        if m:
            data['hostname'] = m.group(1)
            break

    iface = None
    for l in lines:
        m = re.match(r'^interface\s+(\S+)', l)
        if m:
            if iface: data['interfaces'].append(iface)
            iface = {'name': m.group(1), 'ip': '', 'mask': '', 'description': ''}
        elif iface:
            mi = re.match(r'\s+ip address\s+(\S+)\s+(\S+)', l)
            if mi:
                iface['ip'] = mi.group(1); iface['mask'] = mi.group(2)
            md = re.match(r'\s+description\s+(.*)', l)
            if md:
                iface['description'] = md.group(1).strip()
            if re.match(r'^[a-zA-Z#]', l) and not l.startswith(' '):
                data['interfaces'].append(iface); iface = None
    if iface:
        data['interfaces'].append(iface)

    # OSPF
    in_ospf = False; in_area = False; ospf_details = {}
    for l in lines:
        if re.match(r'^ospf', l):
            in_ospf = True; ospf_details = {'router_id': '', 'networks': []}
        elif in_ospf:
            if l.strip() == '#' or re.match(r'^[a-zA-Z]', l):
                if ospf_details.get('networks') or ospf_details.get('router_id'):
                    data['protocols'].append({'type': 'OSPF', 'details': json.dumps(ospf_details)})
                in_ospf = False
            else:
                m = re.match(r'\s+router-id\s+(\S+)', l)
                if m: ospf_details['router_id'] = m.group(1)
                m = re.match(r'\s+network\s+(\S+)\s+(\S+)', l)
                if m: ospf_details['networks'].append({'net': m.group(1)})

    # BGP
    in_bgp = False; bgp_details = {}
    for l in lines:
        if re.match(r'^bgp\s+\d+', l):
            in_bgp = True; bgp_details = {'as': re.search(r'\d+', l).group(), 'neighbors': []}
        elif in_bgp:
            if l.strip() == '#' or re.match(r'^[a-zA-Z]', l):
                data['protocols'].append({'type': 'BGP', 'details': json.dumps(bgp_details)})
                in_bgp = False
            else:
                m = re.match(r'\s+peer\s+(\S+)\s+as-number\s+(\S+)', l)
                if m: bgp_details['neighbors'].append({'ip': m.group(1), 'as': m.group(2)})

    # ACL
    in_acl = False; acl_name = ''
    for l in lines:
        m = re.match(r'^acl number\s+(\S+)', l)
        if m:
            in_acl = True; acl_name = f"ACL {m.group(1)}"
        elif in_acl:
            if l.strip() == '#' or re.match(r'^[a-zA-Z]', l):
                in_acl = False
            else:
                mr = re.match(r'\s+rule\s+(.*)', l)
                if mr:
                    data['acl_rules'].append({'acl_name': acl_name, 'rule': mr.group(1).strip()})

    return data

def parse_juniper(content):
    data = {'hostname': '', 'interfaces': [], 'protocols': [], 'acl_rules': []}

    m = re.search(r'host-name\s+(\S+);', content)
    if m: data['hostname'] = m.group(1)

    # Interfaces
    iface_blocks = re.finditer(
        r'(\S[\w/.-]+)\s*\{[^}]*unit\s+0\s*\{[^}]*family inet\s*\{[^}]*address\s+(\S+);',
        content, re.DOTALL)
    for b in iface_blocks:
        name = b.group(1).strip()
        cidr = b.group(2).strip()
        ip, prefix = cidr.split('/') if '/' in cidr else (cidr, '')
        mask = str(ipaddress.IPv4Network(f'0.0.0.0/{prefix}', strict=False).netmask) if prefix else ''
        desc_m = re.search(r'description\s+"([^"]+)";', b.group(0))
        desc = desc_m.group(1) if desc_m else ''
        data['interfaces'].append({'name': name, 'ip': ip, 'mask': mask, 'description': desc})

    # BGP
    bgp_block = re.search(r'bgp\s*\{(.*?)\}', content, re.DOTALL)
    if bgp_block:
        bgp_details = {'neighbors': []}
        for nb in re.finditer(r'neighbor\s+(\S+);', bgp_block.group(1)):
            bgp_details['neighbors'].append({'ip': nb.group(1)})
        m_local = re.search(r'local-address\s+(\S+);', bgp_block.group(1))
        if m_local: bgp_details['router_id'] = m_local.group(1)
        data['protocols'].append({'type': 'BGP', 'details': json.dumps(bgp_details)})

    # OSPF
    ospf_block = re.search(r'ospf\s*\{(.*?)\}', content, re.DOTALL)
    if ospf_block:
        ospf_details = {'networks': []}
        for nb in re.finditer(r'network\s+(\S+);', ospf_block.group(1)):
            ospf_details['networks'].append({'net': nb.group(1)})
        data['protocols'].append({'type': 'OSPF', 'details': json.dumps(ospf_details)})

    # Policy as ACL equivalent
    for pm in re.finditer(r'policy-statement\s+(\S+)\s*\{(.*?)\}', content, re.DOTALL):
        data['acl_rules'].append({
            'acl_name': pm.group(1),
            'rule': re.sub(r'\s+', ' ', pm.group(2)).strip()
        })

    return data

# ─── Validation ───────────────────────────────────────────────────────────────

def validate_devices():
    devices = Device.query.all()
    # Clear old validations
    Validation.query.delete()
    db.session.commit()

    all_subnets = {}  # subnet -> device hostname

    for dev in devices:
        ifaces = Interface.query.filter_by(device_id=dev.id).all()

        # 1. Loopback check
        has_lo = any(
            re.search(r'(loopback|lo0)', i.name, re.I) for i in ifaces
        )
        db.session.add(Validation(
            device_id=dev.id,
            check='Has Loopback0',
            status='PASS' if has_lo else 'FAIL',
            message='Loopback0 found' if has_lo else 'Missing Loopback0 interface'
        ))

        # 2. Subnet overlap (collect)
        for iface in ifaces:
            if iface.ip and iface.mask:
                try:
                    net = ipaddress.IPv4Network(f'{iface.ip}/{iface.mask}', strict=False)
                    key = str(net)
                    if key in all_subnets:
                        db.session.add(Validation(
                            device_id=dev.id,
                            check='IP Subnet Overlap',
                            status='FAIL',
                            message=f'Subnet {key} on {iface.name} overlaps with {all_subnets[key]}'
                        ))
                    else:
                        all_subnets[key] = f'{dev.hostname}/{iface.name}'
                        db.session.add(Validation(
                            device_id=dev.id,
                            check='IP Subnet Overlap',
                            status='PASS',
                            message=f'Subnet {key} ({iface.name}) is unique'
                        ))
                except Exception:
                    pass

    # 3. OSPF area consistency
    ospf_areas = {}
    for dev in devices:
        protos = Protocol.query.filter_by(device_id=dev.id, type='OSPF').all()
        for p in protos:
            d = json.loads(p.details)
            for net in d.get('networks', []):
                area = net.get('area', 'unknown')
                ospf_areas.setdefault(area, []).append(dev.hostname)

    if ospf_areas:
        all_ospf_devs = Device.query.filter(
            Device.id.in_([p.device_id for p in Protocol.query.filter_by(type='OSPF').all()])
        ).all()
        main_area = max(ospf_areas, key=lambda a: len(ospf_areas[a]))
        for dev in all_ospf_devs:
            protos = Protocol.query.filter_by(device_id=dev.id, type='OSPF').all()
            areas_used = set()
            for p in protos:
                d = json.loads(p.details)
                for net in d.get('networks', []):
                    areas_used.add(net.get('area', 'unknown'))
            consistent = main_area in areas_used
            db.session.add(Validation(
                device_id=dev.id,
                check='OSPF Area Consistency',
                status='PASS' if consistent else 'FAIL',
                message=f'Using area {", ".join(areas_used)}' if consistent else f'Area mismatch: {areas_used}'
            ))

    db.session.commit()

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('upload'))

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    message = None
    if request.method == 'POST':
        files = request.files.getlist('configs')
        if not files or all(f.filename == '' for f in files):
            message = ('error', 'No files selected.')
        else:
            count = 0
            for f in files:
                if f.filename == '':
                    continue
                content = f.read().decode('utf-8', errors='replace')
                vendor = detect_vendor(content, f.filename)

                if vendor == 'Cisco':
                    parsed = parse_cisco(content)
                elif vendor == 'Huawei':
                    parsed = parse_huawei(content)
                else:
                    parsed = parse_juniper(content)

                # Remove existing device with same filename
                existing = Device.query.filter_by(filename=f.filename).first()
                if existing:
                    db.session.delete(existing)
                    db.session.commit()

                dev = Device(hostname=parsed['hostname'], vendor=vendor, filename=f.filename)
                db.session.add(dev)
                db.session.flush()

                for i in parsed['interfaces']:
                    db.session.add(Interface(device_id=dev.id, **i))
                for p in parsed['protocols']:
                    db.session.add(Protocol(device_id=dev.id, **p))
                for a in parsed['acl_rules']:
                    db.session.add(ACLRule(device_id=dev.id, **a))

                db.session.commit()
                count += 1

            validate_devices()
            message = ('success', f'{count} file(s) uploaded and parsed successfully.')

    devices = Device.query.order_by(Device.uploaded_at.desc()).all()
    return render_template('upload.html', message=message, devices=devices)

@app.route('/dashboard')
def dashboard():
    devices = Device.query.all()
    validations = Validation.query.all()

    # Validation summary
    val_summary = []
    for dev in devices:
        dev_vals = [v for v in validations if v.device_id == dev.id]
        passes = sum(1 for v in dev_vals if v.status == 'PASS')
        fails  = sum(1 for v in dev_vals if v.status == 'FAIL')
        val_summary.append({
            'hostname': dev.hostname,
            'vendor': dev.vendor,
            'passes': passes,
            'fails': fails,
            'checks': dev_vals
        })

    # Protocol counts for pie chart
    ospf_count = Protocol.query.filter_by(type='OSPF').count()
    bgp_count  = Protocol.query.filter_by(type='BGP').count()

    # Interface counts per device for bar chart
    iface_data = []
    for dev in devices:
        cnt = Interface.query.filter_by(device_id=dev.id).count()
        iface_data.append({'hostname': dev.hostname, 'count': cnt})

    return render_template('dashboard.html',
        devices=devices,
        val_summary=val_summary,
        ospf_count=ospf_count,
        bgp_count=bgp_count,
        iface_data=iface_data,
        total_pass=sum(1 for v in validations if v.status=='PASS'),
        total_fail=sum(1 for v in validations if v.status=='FAIL'),
    )

@app.route('/export/<fmt>')
def export(fmt):
    devices = Device.query.all()
    rows = []
    for dev in devices:
        ifaces = Interface.query.filter_by(device_id=dev.id).all()
        protos = Protocol.query.filter_by(device_id=dev.id).all()
        acls   = ACLRule.query.filter_by(device_id=dev.id).all()
        vals   = Validation.query.filter_by(device_id=dev.id).all()

        for i in ifaces:
            rows.append({
                'Hostname': dev.hostname, 'Vendor': dev.vendor,
                'Type': 'Interface', 'Name': i.name,
                'IP': i.ip, 'Mask': i.mask, 'Details': i.description
            })
        for p in protos:
            rows.append({
                'Hostname': dev.hostname, 'Vendor': dev.vendor,
                'Type': f'Protocol ({p.type})', 'Name': p.type,
                'IP': '', 'Mask': '', 'Details': p.details
            })
        for a in acls:
            rows.append({
                'Hostname': dev.hostname, 'Vendor': dev.vendor,
                'Type': 'ACL', 'Name': a.acl_name,
                'IP': '', 'Mask': '', 'Details': a.rule
            })
        for v in vals:
            rows.append({
                'Hostname': dev.hostname, 'Vendor': dev.vendor,
                'Type': 'Validation', 'Name': v.check,
                'IP': v.status, 'Mask': '', 'Details': v.message
            })

    if fmt == 'csv':
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=['Hostname','Vendor','Type','Name','IP','Mask','Details'])
        writer.writeheader()
        writer.writerows(rows)
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            mimetype='text/csv',
            as_attachment=True,
            download_name='network_audit.csv'
        )
    elif fmt == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = 'Network Audit'
        headers = ['Hostname','Vendor','Type','Name','IP','Mask','Details']
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(headers, 1):
                ws.cell(row=ri, column=ci, value=row[key])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='network_audit.xlsx')

    return 'Unknown format', 400

@app.route('/api/devices')
def api_devices():
    devices = Device.query.all()
    result = []
    for dev in devices:
        ifaces = Interface.query.filter_by(device_id=dev.id).all()
        protos = Protocol.query.filter_by(device_id=dev.id).all()
        result.append({
            'id': dev.id, 'hostname': dev.hostname, 'vendor': dev.vendor,
            'interfaces': [{'name': i.name, 'ip': i.ip, 'mask': i.mask} for i in ifaces],
            'protocols': [{'type': p.type} for p in protos]
        })
    return jsonify(result)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5000)
